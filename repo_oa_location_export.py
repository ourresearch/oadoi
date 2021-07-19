import csv
import gzip
import json
import os
from datetime import datetime

from google.cloud import bigquery
from openpyxl import Workbook
from sendgrid.helpers.mail import TrackingSettings, ClickTracking
from sqlalchemy import text

from app import db
from app import logger
from emailer import create_email, send
from endpoint import Endpoint
from repo_oa_location_export_request import RepoOALocationExportRequest

MAX_RESULT_ROWS = 200000

def _bigquery_query_result(endpoint_id):
    _setup_bigquery_creds()
    client = bigquery.Client()

    query_text = '''
        select
            api_live.doi as published_doi,
            api_live.title as published_title,
            replace(replace(oa_location.evidence, 'oa repository (via OAI-PMH ', ''), ' match)', '') as match_evidence,
            coalesce(pmh_record.pmh_id, pmh_record.id) as pmh_record_id,
            pmh_record.title as pmh_record_title,
            pmh_record.doi as pmh_record_doi,
            oa_location.url_for_landing_page,
            oa_location.url_for_pdf,
            oa_location.version,
            oa_location.license
        from
            unpaywall.api_live,
            unnest(oa_locations) as oa_location
        join
            pmh.pmh_record on oa_location.endpoint_id = pmh_record.endpoint_id and (oa_location.pmh_id = pmh_record.pmh_id or oa_location.pmh_id = pmh_record.id)
        where
            oa_location.endpoint_id = @endpoint_id
        order by pmh_record_id, published_doi, oa_location.url_for_landing_page, version
        ;
    '''

    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter("endpoint_id", "STRING", endpoint_id)]
    )

    query_job = client.query(query_text, job_config=job_config, location="US")
    return [dict(row) for row in query_job.result(max_results=MAX_RESULT_ROWS)]


# export GOOGLE_SHEETS_CREDS_JSON=`heroku config:get GOOGLE_SHEETS_CREDS_JSON`
def _setup_bigquery_creds():
    # get creds and save in a temp file because google needs it like this
    json_creds = os.getenv("GOOGLE_SHEETS_CREDS_JSON")
    creds_dict = json.loads(json_creds)
    creds_dict["private_key"] = creds_dict["private_key"].replace("\\\\n", "\n")
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "google_application_credentials.json"
    with open('google_application_credentials.json', 'w') as outfile:
        json.dump(creds_dict, outfile)


def _send_result_email(export_request, result_rows):
    endpoint = Endpoint.query.get(export_request.endpoint_id)

    files = []

    if result_rows:
        fieldnames = [
            'published_doi',
            'published_title',
            'match_evidence',
            'pmh_record_id',
            'pmh_record_title',
            'pmh_record_doi',
            'url_for_landing_page',
            'url_for_pdf',
            'version',
            'license',
        ]

        csv_name = 'oa_locations.csv.gz'
        with gzip.open(csv_name, 'wt') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, dialect='excel')
            writer.writeheader()
            for my_dict in result_rows:
                writer.writerow(my_dict)

        files.append(csv_name)

        excel_name = 'oa_locations.xlsx'
        book = Workbook()
        sheet = book.worksheets[0]
        sheet.title = "oa_locations"

        for col_idx, field_name in enumerate(fieldnames):
            sheet.cell(column=col_idx + 1, row=1, value=field_name)

        for row_idx, row in enumerate(result_rows):
            for col_idx, field_name in enumerate(fieldnames):
                sheet.cell(column=col_idx + 1, row=row_idx + 2, value=row[field_name])

        book.save(filename=excel_name)

        files.append(excel_name)

    if endpoint and endpoint.repo and endpoint.repo.repository_name and endpoint.repo.institution_name:
        repo_description = f'{endpoint.repo.repository_name} at {endpoint.repo.institution_name}'
    elif endpoint and endpoint.repo and endpoint.repo.repository_name:
        repo_description = endpoint.repo.repository_name
    else:
        repo_description = f'Repository ID {export_request.endpoint_id}'

    email = create_email(
        export_request.email,
        f'Unpaywall Open Access report for {repo_description}',
        'repo_oa_location_export',
        {
            "data": {
                "repo_description": repo_description,
                "endpoint_id": export_request.endpoint_id,
                "has_results": bool(result_rows)
            }
        },
        files
    )

    tracking_settings = TrackingSettings()
    tracking_settings.click_tracking = ClickTracking(False, False)
    email.tracking_settings = tracking_settings

    send(email, for_real=True)


def _run():
    pending_request_query = '''
            with pending_requests as (
                select id from repo_oa_location_export_request
                where finished is null and (started is null or started < now() - interval '1 hour')
                for update skip locked
            )
            update repo_oa_location_export_request update_rows
            set started=now()
            from pending_requests
            where update_rows.id = pending_requests.id
            returning pending_requests.id
        '''

    pending_request_ids = [
        row[0] for row in
        db.engine.execute(text(pending_request_query).execution_options(autocommit=True)).fetchall()
    ]

    pending_requests = db.session.query(RepoOALocationExportRequest).filter(
        RepoOALocationExportRequest.id.in_(pending_request_ids)
    ).all()

    for pending_request in pending_requests:
        logger.info(f'processing export request {pending_request}')
        try:
            pending_request.tries += 1
            results = _bigquery_query_result(pending_request.endpoint_id)
            _send_result_email(pending_request, results)
            pending_request.finished = datetime.utcnow()
            pending_request.success = True
        except Exception as e:
            pending_request.success = False
            pending_request.error = str(e)
            if pending_request.tries >= 3:
                pending_request.finished = datetime.utcnow()

        pending_request.started = None
        db.session.commit()


if __name__ == "__main__":
    _run()
