import json
import os
import re
import sys
from collections import defaultdict, OrderedDict
from datetime import date, datetime, timedelta
from time import time

import boto
import redis
import unicodecsv
from flask import Response
from flask import abort
from flask import current_app
from flask import g
from flask import jsonify
from flask import make_response
from flask import redirect
from flask import render_template
from flask import request
from flask import url_for
from redis import WatchError
from openpyxl import Workbook
from sqlalchemy import sql
from sqlalchemy.orm import raiseload
from urllib.parse import quote

import journal_export
import pub
import repository
from app import app
from app import db
from app import logger
from changefile import DAILY_FEED, WEEKLY_FEED
from changefile import get_changefile_dicts
from changefile import get_file_from_bucket
from changefile import valid_changefile_api_keys
from emailer import create_email
from emailer import send
from endpoint import Endpoint
from endpoint import lookup_endpoint_by_pmh_url
from monitoring.error_reporting import handle_papertrail_alert
from page import LANDING_PAGE_ARCHIVE_BUCKET as REPO_PAGE_ARCHIVE_BUCKET
from page import PageNew
from pmh_record import PmhRecord
from put_repo_requests_in_db import add_endpoint
from recordthresher.pubmed import PubmedRaw
from repo_oa_location_export_request import RepoOALocationExportRequest
from repo_pulse import BqRepoPulse
from repo_request import RepoRequest
from repository import Repository
from search import autocomplete_phrases
from search import fulltext_search_title
from snapshot import get_daily_snapshot_key
from static_api_response import StaticAPIResponse
from util import NoDoiException
from util import clean_doi, normalize_doi
from util import elapsed
from util import restart_dynos
from util import str_to_bool


def json_dumper(obj):
    """
    if the obj has a to_dict() function we've implemented, uses it to get dict.
    from http://stackoverflow.com/a/28174796
    """
    try:
        return obj.to_dict()
    except AttributeError:
        return obj.__dict__


def json_resp(thing):
    json_str = json.dumps(thing, sort_keys=True, default=json_dumper, indent=4)

    if request.path.endswith(".json") and (os.getenv("FLASK_DEBUG", False) == "True"):
        logger.info("rendering output through debug_api.html template")
        resp = make_response(render_template(
            'debug_api.html',
            data=json_str))
        resp.mimetype = "text/html"
    else:
        resp = make_response(json_str, 200)
        resp.mimetype = "application/json"
    return resp


def abort_json(status_code, msg):
    body_dict = {
        "HTTP_status_code": status_code,
        "message": msg,
        "error": True
    }
    resp_string = json.dumps(body_dict, sort_keys=True, indent=4)
    resp = make_response(resp_string, status_code)
    resp.mimetype = "application/json"
    abort(resp)


def log_request(resp):
    log_dict = {}
    if resp.status_code == 200:
        if request.endpoint == "get_doi_endpoint":
            try:
                results = json.loads(resp.get_data())["results"][0]
            except Exception:
                # don't bother logging if no results
                return
            oa_color = results["oa_color"]
            if not oa_color:
                oa_color = "gray"
            if oa_color == "green":
                host_type = "repository"
            elif oa_color == "gray":
                host_type = None
            else:
                host_type = "publisher"

            log_dict = {
                "doi": results["doi"],
                "email": request.args.get("email", "no_email_given"),
                "year": results.get("year", None),
                "publisher": None,
                "is_oa": oa_color != "gray",
                "host_type": host_type,
                "license": results.get("license", None),
                "journal_is_oa": None
            }
        elif request.endpoint == "get_doi_endpoint_v2":
            try:
                results = json.loads(resp.get_data())
                best_oa_location = results.get("best_oa_location", None)
            except Exception:
                # don't bother logging if no results
                return
            host_type = None
            license = None
            if best_oa_location:
                host_type = best_oa_location.get("host_type", None)
                license = best_oa_location.get("license", None)
            log_dict = {
                "doi": results["doi"],
                "email": request.args.get("email", "no_email_given"),
                "api_key": request.args.get("api_key", "no_api_key_given"),
                "year": results.get("year", None),
                "publisher": results.get("publisher", None),
                "is_oa": results.get("is_oa", None),
                "host_type": host_type,
                "license": license,
                "journal_is_oa": results.get("journal_is_oa", None)
            }

    if log_dict:
        logger.info("logthis: {}".format(json.dumps(log_dict)))


@app.after_request
def after_request_stuff(resp):

    #support CORS
    resp.headers['Access-Control-Allow-Origin'] = "*"
    resp.headers['Access-Control-Allow-Methods'] = "POST, GET, OPTIONS, PUT, DELETE, PATCH"
    resp.headers['Access-Control-Allow-Headers'] = "origin, content-type, accept, x-requested-with"

    # remove session
    db.session.remove()

    # without this jason's heroku local buffers forever
    sys.stdout.flush()

    # log request for analytics
    log_request(resp)

    return resp


@app.before_request
def stuff_before_request():
    if request.endpoint in ["get_doi_endpoint_v2", "get_doi_endpoint", "get_search_query"]:
        email = request.args.get("email", None)
        api_key = request.args.get("api_key", None)

        if api_key:
            if api_key not in valid_changefile_api_keys():
                abort_json(403, "Invalid api_key")
        else:
            if not email:
                abort_json(422, "Email address required in API call, see http://unpaywall.org/products/api")
            if (email.endswith("example.com") and email != "unpaywall_01@example.com") or email == "YOUR_EMAIL":
                abort_json(422, "Please use your own email address in API calls. See http://unpaywall.org/products/api")

            placeholder_emails = [
                (r'^email@.*email\.com$', date(2021, 2, 18)),
                (r'^your@email\.org$', date(2021, 2, 18)),
                (r'^[^@]+$', date(2021, 2, 18)),
                (r'^enter-your-email@your-institution-domain\.edu$', date(2021, 2, 18)),
                (r'randomEmail', date(1970, 1, 1)),
                (r'^your.email@e?mail\.*', date(1970, 1, 1)),
            ]

            for placeholder_email, block_date in placeholder_emails:
                if re.search(placeholder_email, email, re.IGNORECASE):
                    if (
                        date.today() == block_date or
                        date.today() == block_date + timedelta(days=7) or
                        date.today() >= block_date + timedelta(days=14)
                    ):
                        abort_json(
                            422,
                            "Please use your own email address in API calls. See http://unpaywall.org/products/api"
                        )

            ip = get_ip()

            try:
                too_many_emails = too_many_emails_per_ip(ip, email)
            except Exception as e:
                logger.exception(f'error in email rate limiting: {e}')
                too_many_emails = False

            if too_many_emails:
                # just testing for now
                logger.info(f'too many emails for {ip}, {email}')

    if get_ip() in [
        "35.200.160.130", "45.249.247.101",  "137.120.7.33",
        "52.56.108.147",  "193.137.134.252", "130.225.74.231"
    ]:
        abort_json(429, "History of API use exceeding rate limits, please email support@unpaywall.org for other data access options, including free full database dump.")

    g.request_start_time = time()
    g.hybrid = 'hybrid' in list(request.args.keys())
    if g.hybrid:
        logger.info("GOT HYBRID PARAM so will run with hybrid.")

    # don't redirect http api in some cases
    if request.url.startswith("http://api."):
        return
    if "staging" in request.url or "localhost" in request.url:
        return

    # redirect everything else to https.
    new_url = None
    try:
        if request.headers["X-Forwarded-Proto"] == "https":
            pass
        elif "http://" in request.url:
            new_url = request.url.replace("http://", "https://")
    except KeyError:
        # logger.info(u"There's no X-Forwarded-Proto header; assuming localhost, serving http.")
        pass

    # redirect to naked domain from www
    if request.url.startswith("https://www.oadoi.org"):
        new_url = request.url.replace(
            "https://www.oadoi.org",
            "https://oadoi.org"
        )
        logger.info("URL starts with www; redirecting to " + new_url)

    if new_url:
        return redirect(new_url, 301)  # permanent


def too_many_emails_per_ip(ip, email):
    redis_client = get_redis_client()

    if not redis_client:
        return False

    max_emails_per_ip = 20
    window_seconds = 300

    redis_key = f'v2-api-ip-emails:{ip}'

    redis_client.zremrangebyscore(redis_key, 0, time() - window_seconds)
    redis_client.zadd(redis_key, {email: time()})
    redis_client.expire(redis_key, window_seconds)

    emails_per_ip = redis_client.zcard(redis_key)

    return emails_per_ip > max_emails_per_ip


_redis_client = None
_redis_init = False


def get_redis_client():
    global _redis_client, _redis_init

    if not _redis_init:
        try:
            _redis_client = redis.from_url(os.environ.get("REDIS_URL"))
        except Exception as e:
            logger.exception(f'failed creating redis client: {e}')

        _redis_init = True

    return _redis_client

# convenience function because we do this in multiple places
def get_multiple_pubs_response():
    is_person_who_is_making_too_many_requests = False

    biblios = []
    body = request.json
    if "dois" in body:
        if len(body["dois"]) > 25:
            abort_json(413, "max number of DOIs is 25")
        if len(body["dois"]) > 1:
            is_person_who_is_making_too_many_requests = True
        for doi in body["dois"]:
            biblios += [{"doi": doi}]
            if "jama" in doi:
                is_person_who_is_making_too_many_requests = True

    elif "biblios" in body:
        for biblio in body["biblios"]:
            biblios += [biblio]

        if len(body["biblios"]) > 1:
            is_person_who_is_making_too_many_requests = True

    logger.info("in get_multiple_pubs_response with {}".format(biblios))

    run_with_hybrid = g.hybrid
    if is_person_who_is_making_too_many_requests:
        logger.info("is_person_who_is_making_too_many_requests, so returning 429")
        abort_json(429, "sorry, you are calling us too quickly.  Please email support@unpaywall.org so we can figure out a good way to get you the data you are looking for.")
    pubs = pub.get_pubs_from_biblio(biblios, run_with_hybrid)
    return pubs


def get_pub_from_doi(doi):
    run_with_hybrid = g.hybrid
    skip_all_hybrid = "skip_all_hybrid" in request.args
    try:
        my_pub = pub.get_pub_from_biblio({"doi": doi},
                                         run_with_hybrid=run_with_hybrid,
                                         skip_all_hybrid=skip_all_hybrid
                                         )
    except NoDoiException:
        msg = "'{}' isn't in Unpaywall. ".format(doi)
        if re.search(r'^10/[a-zA-Z0-9]+', doi):
            msg += 'shortDOIs are not currently supported. '
        msg += 'See https://support.unpaywall.org/a/solutions/articles/44001900286'
        raise NoDoiException(msg)
    return my_pub

@app.route("/repo_pulse/endpoint/institution/<repo_name>", methods=["GET"])
def get_repo_pulse_search_endpoint(repo_name):
    my_repo = Repository.query.filter(Repository.institution_name.ilike("%{}%".format(repo_name))).first()
    my_endpoint = my_repo.endpoints[0]
    endpoint_id = my_endpoint.id
    return get_repo_pulse_endpoint(endpoint_id)


@app.route("/repo_pulse/endpoint/<endpoint_id>/pmh/recent", methods=["GET"])
def get_repo_pulse_endpoint_pmh_recent(endpoint_id):
    version_filter = request.args.get("version", None)
    if version_filter:
        rows = PageNew.query\
            .filter(PageNew.endpoint_id==endpoint_id, PageNew.scrape_version==version_filter)\
            .order_by(PageNew.record_timestamp.desc())\
            .limit(100)
        # deduplicate, because they don't care about the match type of the pages
        results = [r.to_dict(include_id=False) for r in rows]
        results = [dict(t) for t in {tuple(d.items()) for d in results}]
    else:
        rows = PmhRecord.query.options(raiseload('*')).filter(PmhRecord.endpoint_id==endpoint_id).order_by(PmhRecord.record_timestamp.desc()).limit(100)
        results = [r.to_dict() for r in rows]

    results = sorted(results, key=lambda k: k['oaipmh_record_timestamp'], reverse=True)
    return jsonify({"results": results})


@app.route("/debug/endpoint/<endpoint_id>/pmh/all", methods=["GET"])
def get_debug_endpoint_page_all(endpoint_id):
    version_filter = request.args.get("version", None)
    if version_filter:
        rows = PageNew.query\
            .filter(PageNew.endpoint_id==endpoint_id, PageNew.scrape_version==version_filter)\
            .all()
    else:
        rows = PageNew.query.filter(PageNew.endpoint_id==endpoint_id).all()

    rows_by_pmh_id_version = defaultdict(dict)

    for row in rows:
        if not row.scrape_version:
            continue

        existing_row = rows_by_pmh_id_version[row.bare_pmh_id].get(row.scrape_version, None)

        if not existing_row or (row.doi and not existing_row['doi']):
            rows_by_pmh_id_version[row.bare_pmh_id][row.scrape_version] = row.to_dict(include_id=False)
            rows_by_pmh_id_version[row.bare_pmh_id][row.scrape_version]['doi'] = row.doi
            del rows_by_pmh_id_version[row.bare_pmh_id][row.scrape_version]['version']

    return jsonify({"results": rows_by_pmh_id_version})


@app.route("/repo_pulse/endpoint/<endpoint_id>", methods=["GET"])
def get_repo_pulse_endpoint(endpoint_id):
    my_live_endpoint = Endpoint.query.get(endpoint_id)
    live_results = my_live_endpoint.to_dict_repo_pulse()

    my_repo_pulse = BqRepoPulse.query.get(endpoint_id)
    if my_repo_pulse:
        results = my_repo_pulse.to_dict()
        # override the bq status, to get the most recent
        results["status"]["check0_identify_status"] = live_results["status"]["check0_identify_status"]
        results["status"]["check1_query_status"] = live_results["status"]["check1_query_status"]
    else:
        results = live_results
    return jsonify({"results": results})

@app.route("/repository/endpoint/test/<path:url>", methods=["GET"])
def repo_pulse_test_url(url):
    from endpoint import test_harvest_url
    response = test_harvest_url(url)
    results = {
        "check0_identify_status": response["harvest_identify_response"],
        "check1_query_status": response["harvest_test_recent_dates"],
        "sample_pmh_record": response["sample_pmh_record"]
    }
    return jsonify({"results": results})

@app.route("/data/repo_pulse/status/<path:endpoint_id>", methods=["GET"])
def repo_pulse_status_endpoint_id(endpoint_id):
    my_endpoint = Endpoint.query.filter(Endpoint.id==endpoint_id).first()
    return jsonify({"results": my_endpoint.to_dict_status()})


@app.route("/repo_pulse/<path:query_string>", methods=["GET"])
def repo_pulse_get_endpoint(query_string):
    query_parts = query_string.split(",")
    objs = []
    for query_part in query_parts:
        objs += lookup_endpoint_by_pmh_url(query_part)
    return jsonify({"results": [obj.to_dict() for obj in objs]})

@app.route("/debug/repo/search/<path:query_string>", methods=["GET"])
def debug_repo_endpoint_search(query_string):
    repos = repository.get_raw_repo_meta(query_string)
    endpoints = []
    for repo in repos:
        for endpoint in repo.endpoints:
            endpoints.append(endpoint)
    return jsonify({"results": [obj.to_dict() for obj in endpoints]})


@app.route("/repo_pulse/endpoint/<endpoint_id>/request_oa_locations", methods=["POST"])
def repo_oa_location_request(endpoint_id):
    body = request.json
    email_address = body["email"]

    export_request = RepoOALocationExportRequest.query.filter(
        RepoOALocationExportRequest.email == email_address,
        RepoOALocationExportRequest.endpoint_id == endpoint_id,
        RepoOALocationExportRequest.finished == None
    ).first()

    if not export_request:
        export_request = RepoOALocationExportRequest(
            endpoint_id=endpoint_id,
            requested=datetime.utcnow(),
            email=email_address
        )

        db.session.merge(export_request)
        db.session.commit()

    return jsonify({
        'endpoint_id': export_request.endpoint_id,
        'requested': export_request.requested,
        'email': export_request.email
    })


def get_endpoints_from_query_string(query_string):
    if "," in query_string:
        repo_ids = query_string.split(",")
    else:
        repo_ids = [query_string]
    repos = repository.get_repos_by_ids(repo_ids)
    endpoints = []
    for repo in repos:
        for endpoint in repo.endpoints:
            endpoints.append(endpoint)
    return endpoints

@app.route("/debug/repo/<query_string>", methods=["GET"])
def debug_repo_endpoint(query_string):
    endpoints = get_endpoints_from_query_string(query_string)
    return jsonify({"results": [obj.to_dict() for obj in endpoints]})

@app.route("/debug/repo/<query_string>/examples/closed", methods=["GET"])
def debug_repo_examples_closed(query_string):
    endpoints = get_endpoints_from_query_string(query_string)
    return jsonify({"results": [obj.get_closed_pages() for obj in endpoints]})

@app.route("/debug/repo/<query_string>/examples/open", methods=["GET"])
def debug_repo_examples_open(query_string):
    endpoints = get_endpoints_from_query_string(query_string)
    return jsonify({"results": [obj.get_open_pages() for obj in endpoints]})


@app.route("/data/sources/<query_string>", methods=["GET"])
def sources_endpoint_search(query_string):
    objs = repository.get_sources_data(query_string)
    return jsonify({"results": [obj.to_dict() for obj in objs]})


@app.route("/data/sources.csv", methods=["GET"])
def sources_endpoint_csv():
    objs = repository.get_sources_data()
    data_string = '\n'.join([obj.to_csv_row() for obj in objs])
    data_string = data_string.encode("utf-8")
    output = make_response(data_string)
    output.headers["Content-Disposition"] = "attachment; filename=unpaywall_sources.csv"
    output.headers["Content-type"] = "text/csv; charset=UTF-8"
    return output


@app.route("/data/sources", methods=["GET"])
def sources_endpoint():
    sources = repository.get_sources_data_fast()
    return jsonify({"results": [s.to_dict() for s in sources]})


@app.route("/data/repositories", methods=["GET"])
def repositories_endpoint():
    repository_metadata_objects = repository.get_repository_data()
    return jsonify({"results": [repo_meta.to_dict() for repo_meta in repository_metadata_objects]})


@app.route("/v1/publication/doi/<path:doi>", methods=["GET"])
@app.route("/v1/publication/doi.json/<path:doi>", methods=["GET"])
def get_from_new_doi_endpoint(doi):
    try:
        my_pub = get_pub_from_doi(doi)
        return jsonify({"results": [my_pub.to_dict_v1()]})
    except NoDoiException as e:
        abort_json(404, str(e))


def get_ip():
    # from http://stackoverflow.com/a/12771438/596939
    if request.headers.getlist("X-Forwarded-For"):
       ip = request.headers.getlist("X-Forwarded-For")[0]
    else:
       ip = request.remote_addr
    return ip


def print_ip():
    user_agent = request.headers.get('User-Agent')
    logger.info("calling from IP {ip}. User-Agent is '{user_agent}'.".format(
        ip=get_ip(),
        user_agent=user_agent
    ))


# this is the old way of expressing this endpoint.
# the new way is POST api.oadoi.org/
# you can give it an object that lists DOIs
# you can also give it an object that lists biblios.
# this is undocumented and is just for impactstory use now.
@app.route("/v1/publications", methods=["POST"])
def new_post_publications_endpoint():
    print_ip()
    pubs = get_multiple_pubs_response()
    if not pubs:
        abort_json(500, "something went wrong.  please email support@unpaywall.org and we'll have a look!")
    return jsonify({"results": [p.to_dict_v1() for p in pubs]})



# this endpoint is undocumented for public use, and we don't really use it
# in production either.
# it's just for testing the POST biblio endpoint.
@app.route("/biblios", methods=["GET"])
@app.route("/biblios.json", methods=["GET"])
@app.route("/v1/publication", methods=["GET"])
@app.route("/v1/publication.json", methods=["GET"])
def get_from_biblio_endpoint():
    request_biblio = {}
    for (k, v) in request.args.items():
        request_biblio[k] = v
    run_with_hybrid = g.hybrid
    my_pub = pub.get_pub_from_biblio(request_biblio, run_with_hybrid=run_with_hybrid)
    return json_resp({"results": [my_pub.to_dict()]})




@app.route("/favicon.ico")
def favicon_ico():
    return redirect(url_for("static", filename="img/favicon.ico"))

@app.route("/browser-tools/bookmarklet.js")
def bookmarklet_js():
    base_url = request.url.replace(
        "browser-tools/bookmarklet.js",
        "static/browser-tools/"
    )

    if "localhost:" not in base_url:
        # seems like this shouldn't be necessary. but i think
        # flask's request.url is coming in with http even when
        # we asked for https on the server. weird.
        base_url = base_url.replace("http://", "https://")

    rendered = render_template(
        "browser-tools/bookmarklet.js",
        base_url=base_url
    )
    resp = make_response(rendered, 200)
    resp.mimetype = "application/javascript"
    return resp



@app.route('/', methods=["GET", "POST"])
def base_endpoint():
    return jsonify({
        "version": "1.3.0",
        "documentation_url": "https://unpaywall.org/data",
        "msg": "Don't panic"
    })

@app.route('/v2', methods=["GET", "POST"])
@app.route('/v2/', methods=["GET", "POST"])
def base_endpoint_v2():
    return jsonify({
        "version": "2.0.1",
        "documentation_url": "https://unpaywall.org/api/v2",
        "msg": "Don't panic"
    })

@app.route("/<path:doi>", methods=["GET"])
def get_doi_endpoint(doi):
    return get_doi_endpoint_v2(doi)

@app.route("/v2/<path:doi>", methods=["GET"])
def get_doi_endpoint_v2(doi):
    # the GET api endpoint (returns json data)
    try:
        my_pub = get_pub_from_doi(doi)
        answer = my_pub.to_dict_v2()
    except NoDoiException as e:
        answer = {}
        normalized_doi = normalize_doi(doi, return_none_if_error=True)
        if normalized_doi:
            static_response = StaticAPIResponse.query.get(doi)
            if static_response:
                answer = OrderedDict([
                    (key, static_response.response_jsonb.get(key, None)) for key in pub.Pub.dict_v2_fields().keys()
                ])
        if not answer:
            abort_json(404, str(e))

    indent = None
    if current_app.config['JSONIFY_PRETTYPRINT_REGULAR'] and not request.is_xhr:
        indent = 2

    return current_app.response_class(json.dumps(answer, indent=indent), mimetype='application/json')


@app.route("/v2/dois", methods=["POST"])
def simple_query_tool():
    body = request.json
    dirty_dois_list = [d for d in body["dois"] if d]

    # look up normalized dois
    normalized_dois = [n for n in [normalize_doi(d, return_none_if_error=True) for d in dirty_dois_list] if n]
    q = db.session.query(pub.Pub.response_jsonb).filter(pub.Pub.id.in_(set(normalized_dois)))
    normalized_doi_responses = dict([(row[0]['doi'], row[0]) for row in q.all() if row[0]])

    # look up cleaned dois
    cleaned_dois = [c for c in [clean_doi(d, return_none_if_error=True) for d in dirty_dois_list] if c]
    q = db.session.query(pub.Pub.response_jsonb).filter(pub.Pub.id.in_(set(cleaned_dois)))
    cleaned_doi_responses = dict([(row[0]['doi'], row[0]) for row in q.all() if row[0]])

    responses = [
        normalized_doi_responses.get(normalize_doi(d, return_none_if_error=True), None)
        or cleaned_doi_responses.get(clean_doi(d, return_none_if_error=True), None)
        or pub.build_new_pub(d, None).to_dict_v2()
        for d in dirty_dois_list
    ]

    formats = body.get("formats", []) or ["jsonl", "csv"]
    files = []

    if "jsonl" in formats:
        # save jsonl
        with open("output.jsonl", 'w') as f:
            for response_jsonb in responses:
                f.write(json.dumps(response_jsonb, sort_keys=True))
                f.write("\n")
        files.append("output.jsonl")

    csv_dicts = [pub.csv_dict_from_response_dict(my_dict) for my_dict in responses]
    csv_dicts = [my_dict for my_dict in csv_dicts if my_dict]
    fieldnames = sorted(csv_dicts[0].keys())
    fieldnames = ["doi"] + [name for name in fieldnames if name != "doi"]

    if "csv" in formats:
        # save csv
        with open("output.csv", 'wb') as f:
            writer = unicodecsv.DictWriter(f, fieldnames=fieldnames, dialect='excel')
            writer.writeheader()
            for my_dict in csv_dicts:
                writer.writerow(my_dict)
        files.append("output.csv")

    if "xlsx" in formats:
        book = Workbook()
        sheet = book.worksheets[0]
        sheet.title = "results"

        for col_idx, field_name in enumerate(fieldnames):
            sheet.cell(column=col_idx+1, row=1, value=field_name)

        for row_idx, row in enumerate(csv_dicts):
            for col_idx, field_name in enumerate(fieldnames):
                sheet.cell(column=col_idx+1, row=row_idx+2, value=row[field_name])

        book.save(filename="output.xlsx")
        files.append("output.xlsx")

    # prep email
    email_address = body["email"]
    email = create_email(email_address,
                 "Your Unpaywall results",
                 "simple_query_tool",
                 {"profile": {}},
                 files)
    send(email, for_real=True)

    return jsonify({
        "got it": email_address,
        "dois": [r['doi'] for r in responses]
    })


@app.route("/repository", methods=["POST"])
def repository_post_endpoint():
    body = request.json
    repo_request = RepoRequest(**body)

    new_endpoint = add_endpoint(repo_request)
    if not new_endpoint:
        abort_json(422, "missing arguments")

    return jsonify({"response": new_endpoint.to_dict()})


def get_s3_csv_gz(s3_key):
    def generate_file():
        for chunk in s3_key:
            yield chunk

    return Response(generate_file(), headers={
        'Content-Length': s3_key.size,
        'Content-Disposition': 'attachment; filename="{}"'.format(s3_key.name),
        'Content-Type': 'application/gzip',
    })


@app.route("/journals.csv.gz", methods=["GET"])
def get_journals_csv():
    return get_s3_csv_gz(journal_export.get_journal_file_key(journal_export.JOURNAL_FILE))


@app.route("/journal_open_access.csv.gz", methods=["GET"])
def get_journal_open_access():
    return get_s3_csv_gz(journal_export.get_journal_file_key(journal_export.OA_STATS_FILE))


@app.route("/repositories.csv.gz", methods=["GET"])
def get_repository_journal_stats():
    return get_s3_csv_gz(journal_export.get_journal_file_key(journal_export.REPO_FILE))


@app.route("/extension_requests.csv.gz", methods=["GET"])
def get_journal_extension_requests():
    return get_s3_csv_gz(journal_export.get_journal_file_key(journal_export.REQUESTS_FILE))


@app.route("/crossref_issns.csv.gz", methods=["GET"])
def get_crossref_issns():
    return get_s3_csv_gz(journal_export.get_journal_file_key(journal_export.ISSNS_FILE))


@app.route("/feed/changefiles", methods=["GET"])
def get_changefiles():
    # api key is optional here, is just sends back urls that populate with it
    api_key = request.args.get("api_key", "YOUR_API_KEY")
    interval = request.args.get("interval", "week")

    if interval == "week":
        feed = WEEKLY_FEED
    elif interval == "day":
        feed = DAILY_FEED
    else:
        abort_json(401, 'option "interval" must be one of ["day", "week"]')

    resp = get_changefile_dicts(api_key, feed=feed)
    return jsonify({"list": resp})


@app.route("/feed/changefile/<path:filename>", methods=["GET"])
def get_changefile_filename(filename):
    api_key = request.args.get("api_key", None)
    if not api_key:
        abort_json(401, "You must provide an API_KEY")
    if api_key not in valid_changefile_api_keys():
        abort_json(403, "Invalid api_key")

    key = get_file_from_bucket(filename)

    def generate_changefile():
        for chunk in key:
            yield chunk

    return Response(generate_changefile(), content_type="gzip", headers={
        'Content-Length': key.size,
        'Content-Disposition': 'attachment; filename="{}"'.format(key.name),
    })


@app.route("/snapshot", methods=["GET"])
@app.route("/feed/snapshot", methods=["GET"])
def get_snapshot():
    api_key = request.args.get("api_key", None)
    if not api_key:
        abort_json(401, "You must provide an API_KEY")
    if api_key not in valid_changefile_api_keys():
        abort_json(403, "Invalid api_key")

    key = get_daily_snapshot_key()

    if key is None:
        abort_json(404, "no snapshots ready")

    def generate_snapshot():
        for chunk in key:
            yield chunk

    return Response(generate_snapshot(), content_type="gzip", headers={
        'Content-Length': key.size,
        'Content-Disposition': 'attachment; filename="{}"'.format(key.name),
    })


@app.route("/daily-feed/changefile/<path:filename>", methods=["GET"])
def get_daily_changefile_filename(filename):
    api_key = request.args.get("api_key", None)
    if not api_key:
        abort_json(401, "You must provide an API_KEY")
    if api_key not in valid_changefile_api_keys():
        abort_json(403, "Invalid api_key")

    key = get_file_from_bucket(filename, feed=DAILY_FEED)

    def generate_changefile():
        for chunk in key:
            yield chunk

    return Response(generate_changefile(), content_type="gzip", headers={
        'Content-Length': key.size,
        'Content-Disposition': 'attachment; filename="{}"'.format(key.name),
    })


@app.route("/issn_ls", methods=["GET", "POST"])
def get_issnls():
    if request.method == 'GET':
        issns = request.args.get('issns', '').split(',')
    else:
        if request.json and isinstance(request.json.get('issns', None), list):
            issns = request.json.get('issns')
        else:
            abort_json(400, 'send a json object like {"issns": ["0005-0970","1804-6436"]}')

    query = sql.text('select issn, issn_l from journalsdb_issn_to_issn_l where issn = any(:issns)').bindparams(issns=issns)

    issn_l_list = db.engine.execute(query).fetchall()
    issn_l_map = dict([(issn_pair[0], issn_pair[1]) for issn_pair in issn_l_list])

    response = {'issn_ls': [issn_l_map.get(issn, None) for issn in issns]}

    return jsonify(response)


@app.route("/v2/search/", methods=["GET"])
def get_search_query():
    query = request.args.get("query", None)
    is_oa = request.args.get("is_oa", None)
    page = request.args.get("page", None)

    if is_oa is not None:
        try:
            is_oa = str_to_bool(is_oa)
        except ValueError:
            if is_oa == 'null':
                is_oa = None
            else:
                abort_json(400, "is_oa must be 'true' or 'false'")

    if page is not None:
        try:
            page = int(page)
            if page < 1:
                raise ValueError
        except ValueError:
            abort_json(400, "'page' must be a positive integer")
    else:
        page = 1

    if not query:
        abort_json(400, "query parameter is required")

    start_time = time()
    response = fulltext_search_title(query, is_oa, page=page)
    sorted_response = sorted(response, key=lambda k: k['score'], reverse=True)

    for api_response in sorted_response:
        doi = api_response['response']['doi']
        version_suffix = re.findall(r'[./](v\d+)$', doi, re.IGNORECASE)

        if version_suffix:
            title = api_response['response']['title']
            title = '{} ({})'.format(title, version_suffix[0].upper())
            api_response['response']['title'] = title

    elapsed_time = elapsed(start_time, 3)
    return jsonify({"results": sorted_response, "elapsed_seconds": elapsed_time})

@app.route("/search/autocomplete/<path:query>", methods=["GET"])
def get_search_autocomplete_query(query):
    start_time = time()
    response = autocomplete_phrases(query)
    sorted_response = sorted(response, key=lambda k: k['score'], reverse=True)
    elapsed_time = elapsed(start_time, 3)
    return jsonify({"results": sorted_response, "elapsed_seconds": elapsed_time})

@app.route("/admin/restart/<api_key>", methods=["GET"])
def restart_endpoint(api_key):
    print("in restart endpoint")
    if api_key != os.getenv("HEROKU_API_KEY"):
        print("not allowed to reboot in restart_endpoint")
        return jsonify({
            "response": "not allowed to reboot, didn't send right heroku api key"
        })

    dyno_prefix = "run_page."
    restart_dynos("oadoi", dyno_prefix)

    return jsonify({
        "response": "restarted dynos: {}".format(dyno_prefix)
    })


@app.route('/admin/report-error/<api_key>', methods=['GET', 'POST'])
def report_error(api_key):
    if api_key != os.getenv("HEROKU_API_KEY"):
        error = 'wrong heroku API key {} in /admin/report-error/'.format(api_key)
        logger.error(error)
        return make_response(error, 403)

    handle_papertrail_alert(request)

    return jsonify({
        'error-data': request.form
    })


@app.route("/pmh_record_xml/<path:pmh_record_id>", methods=["GET"])
def get_pmh_record_xml(pmh_record_id):
    record = PmhRecord.query.get(pmh_record_id)

    if not record or not record.api_raw:
        return Response('', mimetype='text/xml', status=404)
    else:
        return Response(record.api_raw, mimetype='text/xml')


@app.route("/crossref_api_cache/<path:doi>", methods=["GET"])
def get_crossref_api_json(doi):
    my_pub = pub.Pub.query.get(normalize_doi(doi))

    if not my_pub or not my_pub.crossref_api_raw_new:
        abort_json(404, f"Can't find a crossref API record for {doi}")
    else:
        return jsonify(my_pub.crossref_api_raw_new)


@app.route("/pubmed_xml/<pmid>", methods=["GET"])
def get_pubmed_xml(pmid):
    pubmed_raw = PubmedRaw.query.get(pmid)

    if not pubmed_raw or not pubmed_raw.pubmed_article_xml:
        abort_json(404, f"Can't find a PubMed record for PMID {pmid}")
    else:
        return Response(pubmed_raw.pubmed_article_xml, mimetype='text/xml')


@app.route("/doi_page/<path:doi>", methods=["GET"])
def get_doi_landing_page(doi):
    doi_key = quote(normalize_doi(doi), safe='')

    s3 = boto.connect_s3()
    bucket = s3.get_bucket(pub.LANDING_PAGE_ARCHIVE_BUCKET)
    key = bucket.lookup(doi_key)

    if not key:
        abort_json(404, f"Can't find a landing page archive for {doi}")

    def generate_file():
        for chunk in key:
            yield chunk

    return Response(generate_file(), content_type="gzip", headers={
        'Content-Length': key.size,
        'Content-Disposition': 'attachment; filename="{}.gz"'.format(key.name),
    })


@app.route("/repo_page/<page_id>", methods=["GET"])
def get_repository_page(page_id):
    if repo_page := PageNew.query.get(page_id):
        if repo_page.landing_page_archive_key:
            s3 = boto.connect_s3()
            bucket = s3.get_bucket(REPO_PAGE_ARCHIVE_BUCKET)
            if key := bucket.lookup(repo_page.landing_page_archive_key.key):
                def generate_file():
                    for chunk in key:
                        yield chunk

                return Response(generate_file(), content_type="gzip", headers={
                    'Content-Length': key.size,
                    'Content-Disposition': 'attachment; filename="{}"'.format(key.name),
                })

    abort_json(404, f"Can't find an archive for repo page {page_id}")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=True, threaded=True)
